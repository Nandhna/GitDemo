import openpyxl


class HomePageData:

        test_HomePage_data = [{"firstname":"Rahul Shetty","emailid":"rahulshetty123@gmail.com","password":"Rahul@123","gender":"Male"},{"firstname":"Rahul Anshika","emailid":"rahulanshika@gmail.com","password":"Rahul@456","gender":"Female"}]

        def getTestData(self,test_case_name):
                Dict = {}
                book = openpyxl.load_workbook("C:\Users\NandhanaRajendran\OneDrive - techcarrot.ae\Documents\PYCHARM projects\Python SelfFramework\test_data\PythonDemo.xlsx")
                sheet = book.active
                for i in range(1, sheet.max_row + 1):
                        if sheet.cell(row=i, colum=1).value == test_case_name:

                                for j in range(2, sheet.max_row + 1):
                                        # Dict["lastname"] = "shetty"
                                        Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

                print(Dict)

