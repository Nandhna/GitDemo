import openpyxl
book = openpyxl.load_workbook("C:\\Users\\NandhanaRajendran\\OneDrive - techcarrot.ae\\Documents\PYCHARM projects\\Python SelfFramework\\test_data\\PythonDemo.xlsx")

sheet = book.active
Dict = {}
cell = sheet.cell(row=1,column=2)      # read the excel file
print(cell.value)

sheet.cell(row=1,column=2).value = "Rahul"                #write the excel file
print(sheet.cell(row=1,column=2).value)

print(sheet.max_row)
print(sheet.max_column)

print(sheet['A5'].value)

for i in range(1,sheet.max_row+1):
    if sheet.cell(row=i, colum=1).value == "Testcase2":

        for j in range(2, sheet.max_row+1):
            #Dict["lastname"] = "shetty"
            Dict[sheet.cell(row=1, column=j).value]= sheet.cell(row=i, column=j).value

print(Dict)




#for i in range(1,sheet.max_row+1):                        #to get rows
#   for j in range(1,sheet.max_row+1):                     #to get columns
#    print(sheet.cell(row=i,column=j).value)              #inorder to print every values in rows andd colums


